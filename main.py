import sys
import PyQt5.QtWidgets
import openpyxl as xl
import os
import pathlib
import win32com.client as win32
import ui
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties


def create_main_book(community_name):
    main_book = xl.Workbook()
    main_book.create_sheet(title="Main")
    ws = main_book.get_sheet_by_name("Sheet")
    main_book.remove_sheet(ws)
    ws = main_book.get_sheet_by_name("Main")
    ws.sheet_properties.tabColor = "29cfcc"
    main_book.save(filename=community_name + "_Close.xlsx")
    return main_book


def convert_files(directory):
    dir = directory
    file_list = os.listdir(dir)
    for file in file_list:
        if file.endswith(".xls"):
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            wb_new = excel.Workbooks.Add()
            wb_new.SaveAs(os.path.join(dir, file) + 'x')
            wb_old = excel.Workbooks.Open(os.path.join(dir, file))

            for sheet in wb_old.Sheets:
                wb_old.Worksheets(sheet.Name).Move(Before=wb_new.Worksheets("Sheet1"))

            wb_new.Worksheets('Sheet1').Delete()
            wb_new.Close(True)
            os.remove(directory + file)


def move_data_to_main_file(main_book, community_name, directory):
    file_list = os.listdir(directory)
    for file in file_list:
        if file.endswith(".xlsx"):
            split_name_list = file.split("_")
            main_sheet = main_book.create_sheet(split_name_list[1][:31])
            wb = xl.load_workbook(filename=directory + file)
            ws = wb.active
            row_max = ws.max_row
            column_max = ws.max_column
            for i in range(1, row_max + 1):
                for j in range(1, column_max + 1):
                    source_cell = ws.cell(row=i, column=j)
                    main_sheet.cell(row=i, column=j).value = source_cell.value
        main_book.save(filename=community_name + "_Close.xlsx")


def delinquent_and_prepaid(sheet):
    balances_to_return = dict();

    print("Checking " + sheet.title + " tab")
    print("Checking all account are either prepaid or delinquent and misc income is $0.00")
    row_max = sheet.max_row
    column_max = sheet.max_column
    for i in range(1, row_max + 1):
        # must create a name_row as not all lines have resident name
        if sheet.cell(column=6, row=i) != "":
            name_row = i - 1
        # column 24 has prepaid and 30 has delinquent and 6 has resident name
        x = sheet.cell(column=24, row=i).value
        y = sheet.cell(column=30, row=i).value
        if x is not None and y is not None:
            x = str(x)
            y = str(y)
            x = x.strip("()")
            y = y.strip("()")
            try:
                x = float(x)
                y = float(y)
                if x != 0 and y != 0:
                    print(sheet.cell(column=6, row=name_row).value + "has a prepaid balance of: $" +
                          str(x) + " and a delinquent balance of: $" + str(y))
            except (ValueError, AttributeError, TypeError):
                if x == None:
                    print(sheet.cell(column=6, row=name_row).value + "has a prepaid balance of: $" +
                          " None " + " and a delinquent balance of: $" + str(y))
                elif y == None:
                    print(sheet.cell(column=6, row=name_row).value + "has a prepaid balance of: $" +
                          str(x) + " and a delinquent balance of: $" + " None")
                else:
                    print("OH NO! Value Error when parsing data")

        if sheet.cell(column=10, row=i).value == "Misc. Income":
            x = str(sheet.cell(column=24, row=i + 1).value)
            y = str(sheet.cell(column=30, row=i + 1).value)
            print(str(sheet.cell(column=6, row=name_row + 1).value) + " is a misc account with a  prepaid"
                                                                      " balance of: $" + str(x) + " and a"
                                                                      " delinquent balance of: $" + str(y))
            
        if sheet.cell(column=31, row=i).value == "Net Prepaid:":
            d_and_p_net_prepaid = str(sheet.cell(column=40, row=i).value)
            d_and_p_net_prepaid = d_and_p_net_prepaid.strip("()")
            balances_to_return["d_and_p_net_prepaid"] = float(d_and_p_net_prepaid)
        if sheet.cell(column=28, row=i).value == "Net Delinquent:":
            balances_to_return["d_and_p_net_delinquent"] = float(sheet.cell(column=40, row=i).value)
    print()
    return balances_to_return


def resident_deposit(sheet):
    print("Checking " + sheet.title + " tab")
    print("Checking all accounts have no deposit on hand")
    row_max = sheet.max_row
    column_max = sheet.max_column
    for i in range(1, row_max + 1):
        if sheet.cell(column=3, row=i) != "":
            name_row = i
        # column 11 has total deposit
        total_deposit = sheet.cell(column=11, row=i).value
        if total_deposit == 0:
            continue
        elif total_deposit is None:
            continue
        else:
            try:
                float(total_deposit)
                print(str(sheet.cell(column=3, row=name_row).value) + " has a deposit balance of: $" +
                      str(total_deposit))
            except ValueError:
                continue
    print()


def scheduled_billing(sheet):
    print("Checking " + sheet.title + " tab")
    print("Checking all scheduled billing is present")
    row_max = sheet.max_row
    column_max = sheet.max_column
    for i in range(1, row_max + 1):
        if str(sheet.cell(column=3, row=i).value) == "Total Billing:":
            pass
        elif str(sheet.cell(column=3, row=i).value) == "":
            pass
        elif sheet.cell(column=3, row=i).value is None:
            pass
        else:
            name_row = i
        # column 8 has next months scheduled billing
        # MUST run the report with Next month first
        scheduled_billing = sheet.cell(column=8, row=i).value
        try:
            scheduled_billing = float(scheduled_billing)
            if scheduled_billing != 0.00:
                continue
            elif scheduled_billing is None:
                continue
            elif sheet.cell(column=1, row=i).value == "Totals:":
                continue
            else:
                print(
                    str(sheet.cell(column=3, row=name_row).value) + " has no scheduled billing for next month")
        except (ValueError, TypeError):
            continue
    print()


def resident_balances(sheet):
    balances_to_return = dict();

    print("Checking " + sheet.title + " tab")
    print("Obtaining net prepaid and net delinquent for audit check")
    row_max = sheet.max_row
    column_max = sheet.max_column
    for i in range(1, row_max + 1):
        if str(sheet.cell(column=31, row=i).value) == "Net Prepaid":
            res_bal_prepaid = str(sheet.cell(column=37, row=i).value)
            res_bal_prepaid = res_bal_prepaid.strip("()")
            balances_to_return['res_bal_prepaid'] = float(res_bal_prepaid)
        if str(sheet.cell(column=31, row=i).value) == "Net Delinquent":
            balances_to_return['res_bal_delinquent'] = float(sheet.cell(column=37, row=i).value)

    print()
    return balances_to_return


def complete_month_end_close(book,community_name):
    for sheet in book:
        if sheet.title.startswith("Delinquent"):
            delinquent_and_prepaid_values = delinquent_and_prepaid(sheet)
        elif sheet.title.startswith("Resident Deposit"):
            resident_deposit(sheet)
        elif sheet.title.startswith("Scheduled Billing"):
            scheduled_billing(sheet)
        elif sheet.title.startswith("Resident Balances"):
            resident_balances_values = resident_balances(sheet)

    print(delinquent_and_prepaid_values["d_and_p_net_prepaid"])
    print(resident_balances_values["res_bal_prepaid"])
    print(delinquent_and_prepaid_values["d_and_p_net_delinquent"])
    print(resident_balances_values["res_bal_delinquent"])
    book.save(filename=community_name + "_Close.xlsx")


def main():
    app = PyQt5.QtWidgets.QApplication(sys.argv)
    win = ui.Example()
    win.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()


